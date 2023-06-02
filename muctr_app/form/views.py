from typing import Any, Optional
from django.db import models
from django.views.generic import CreateView, ListView, UpdateView
from .models import Student
from .forms import StudentForm
from django.urls import reverse
import json
from django.contrib import messages
from django.contrib.auth.models import User, auth
from django.shortcuts import render, redirect
from django.http import HttpResponse
from openpyxl import Workbook

def export_to_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Студенты"

    # Define the headers for the worksheet
    headers = ['Логин', 'Имя', 'Фамилия', 'Отчество', 'Дата рождения', 'Страна проживания', 'Город', 'Адрес', 'Гражданство', 'Пол', 'Эл-почта', 'Номер телефона', 'Второй номер телефона', 'Изменение имени', 'Серия номер паспорта', 'Форма обучения', 'Степень', 'Специальность', 'Фото', 'Паспорт', 'Перевод паспорта', 'Аттестат', 'Перевод аттестата', 'Согласия на проверку документов', 'Согласие на поступление']

    # Write the headers to the first row of the worksheet
    row_num = 1
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header

    # Write the data for each object to subsequent rows of the worksheet
    queryset = Student.objects.filter(checked='checked')
    for obj in queryset:
        row_num += 1
        row = [obj.username, obj.name, obj.lastname, obj.middlename, obj.date_of_birth, obj.country, obj.state, obj.address, obj.citizenship, obj.gender, obj.email, obj.changedname, obj.phone, obj.phone_2, obj.passport_series, obj.edu_type, obj.edu_level, obj.edu_field, obj.photo, obj.passport, obj.passport_translation, obj.school_diploma, obj.school_diploma_translation, obj.info_check_agreement, obj.entrance_agreement]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook to the response object
    wb.save(response)

    return response





class FormView(CreateView):
    model = Student
    form_class = StudentForm
    template_name = 'form.html'
    success_url = 'signin'

    def get_success_url(self) -> str:
        user = User.objects.create(
           username = self.object.username,
           email = self.object.email, 
           password = self.object.passport_series
        )
        user.save()
        return 'signin'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        emails = Student.objects.values_list('email')
        users = User.objects.values_list('username')
        email_list = [email[0] for email in emails]
        user_list = [user[0] for user in users] 
        context['emails'] = json.dumps(email_list)
        context['users'] = json.dumps(user_list)
        return context

class RoomView(UpdateView):
    model = Student
    template_name = "room.html"
    form_class = StudentForm
    success_url = 'https://muctr.uz/ru'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        emails = Student.objects.values_list('email')
        users = User.objects.values_list('username')

        email_list = [email[0] for email in emails]
        user_list = [user[0] for user in users] 

        context['emails'] = json.dumps(email_list)
        context['users'] = json.dumps(user_list)


        # get data from checked info model
        username = context['object'].username
        checked_info = Student.objects.get(username=username)

        context['address'] = json.dumps(checked_info.address_desc)
        context['name'] = json.dumps(checked_info.name_desc)
        context['photo'] = json.dumps(checked_info.photo_desc)
        context['lastname'] = json.dumps(checked_info.lastname_desc)
        context['middlename'] = json.dumps(checked_info.middlename_desc)
        context['passport'] = json.dumps(checked_info.passport_desc)
        context['citizenship'] = json.dumps(checked_info.citizenship_desc)
        context['school_diploma'] = json.dumps(checked_info.school_diploma_desc)
        context['passport_series'] = json.dumps(checked_info.passport_series_desc)
        context['entrance_agreement'] = json.dumps(checked_info.entrance_agreement_desc)
        context['info_check_agreement'] = json.dumps(checked_info.info_check_agreement_desc)
        context['passport_translation'] = json.dumps(checked_info.passport_translation_desc)
        context['school_diploma_translation'] = json.dumps(checked_info.school_diploma_translation_desc)



        return context

# def signup(request):
#   if request.method == 'POST':
#     name = request.POST['name']
#     lastname = request.POST['lastname']
#     username = request.POST['username']
#     email = request.POST['email']
#     phone_number = request.POST['phone_number']
#     password = request.POST['password']
#     password2 = request.POST['password2']

#     if password == password2:
#       if User.objects.filter(email=email).exists():
#         messages.info(request, 'Email Taken')
#         return redirect('signup')
#       elif User.objects.filter(username=username).exists():
#         messages.info(request, 'Username Taken')
#         return redirect('signup')
#       else:
#         user = User.objects.create_user(
#           username=username, 
#           email=email,
#           password=password
#         )
#         user.save()

#         customUser = CustomUser.objects.create(
#           username=user,
#           name=name,
#           lastname=lastname,
#           phone_number=phone_number,
#           password=password
#         )
#         customUser.save()

#         user_login = auth.authenticate(username=username, password=password)
#         auth.login(request, user_login)
#         return redirect('form', pk=user.id)
#     else: 
#       messages.info(request, 'Password Not Matching')
#       return redirect('signup') 
#   else:
#     emails = User.objects.values_list('email')
#     usernames = User.objects.values_list('email')

#     context = {}

#     email_list = [email[0] for email in emails]
#     username_list = [username[0] for username in usernames]
#     context['emails'] = json.dumps(email_list)
#     context['usernames'] = json.dumps(username_list)

#     return render(request, 'signup.html', context=context)

def signup(request):
    if request.method == 'POST':
        name = request.POST['name']
        lastname = request.POST['lastname']
        username = request.POST['username']
        email = request.POST['email']
        phone_number = request.POST['phone_number']
        passport_series = request.POST['passport_series']

        exist_user1 = Student.objects.filter(username=username)
        exist_user2 = Student.objects.filter(email=email)
        exist_user3 = Student.objects.filter(passport_series=passport_series)
        print(exist_user1)
        print(exist_user2)
        print(exist_user3)

        if exist_user1.exists() or exist_user2.exists() or exist_user3.exists():
           print("hi")
           if exist_user1 == exist_user2 == exist_user3:
              return redirect('signin')
           else:
              return redirect('signup')

        else:
            user = User.objects.create(
                username = username,

                email = email, 
                password = passport_series
            )
            user.save()

            student = Student.objects.create(
                name=name, 
                lastname=lastname,
                username=username,
                email=email,
                phone=phone_number,
                passport_series=passport_series
            )
            student.save()

            return redirect('signin')
    else:
        context = {}
        usernames = Student.objects.values_list('username')
        passports = Student.objects.values_list('passport_series')
        emails = User.objects.values_list('email')

        username_list = [username[0] for username in usernames]
        passport_list = [passport[0] for passport in passports]
        email_list = [email[0] for email in emails]

        context['emails'] = json.dumps(email_list)
        context['usernames'] = json.dumps(username_list)
        context['passports'] = json.dumps(passport_list)

        return render(request, 'signup.html', context=context)
    


def signin(request):
  if request.method == 'POST':
    username = request.POST.get('username')
    password = request.POST.get('password')

    student = Student.objects.filter(username=username, passport_series=password)[0]

    if student is not None:
      student = Student.objects.get(username=username, passport_series=password)
      pk = student.id
      url = reverse('room', args=[str(pk)])
      return redirect(url)
    else:
      messages.info(request, 'Credentials Invalid')
      return redirect('signin')
  else:
    context = {}
    usernames = Student.objects.values_list('username')
    passports = Student.objects.values_list('passport_series')

    username_list = [username[0] for username in usernames]

    passport_list = [passport[0] for passport in passports]


    context['usernames'] = json.dumps(username_list)
    context['passports'] = json.dumps(passport_list)

    return render(request, 'signin.html', context=context)
  
